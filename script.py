import glob
import os
import random
import re
import subprocess
from time import sleep

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from playwright.sync_api import sync_playwright

ip = 0

#
# Your ProxyCrawl API ke

auth_file = r"C:\Users\m4a1\Documents\auth.txt"

# === CONFIG ===
OPENVPN_PATH = r"C:\Program Files\OpenVPN\bin\openvpn.exe"
CONFIG_DIR = r"C:\Users\m4a1\Documents\ovpn_udp"
current_index = 0
vpn_process = None


ovpn_files = glob.glob(os.path.join(CONFIG_DIR, "*.ovpn"))


def rotate_ip():
    global current_index, vpn_process
    if vpn_process:
        vpn_process.terminate()
        vpn_process.wait()
    subprocess.run(["powershell", "-Command", "Disable-NetAdapter -Name 'OpenVPN TAP-Windows6' -Confirm:$false"],
                   shell=True)
    sleep(2)
    subprocess.run(["powershell", "-Command", "Enable-NetAdapter -Name 'OpenVPN TAP-Windows6' -Confirm:$false"],
                   shell=True)
    sleep(2)
    current_index = current_index + 1
    if current_index == len(ovpn_files):
        print("out of serbers idiot")
        exit(1)
    print(f"Connecting to {ovpn_files[current_index]}...")
    vpn_process = subprocess.Popen([
        OPENVPN_PATH,
        "--config", ovpn_files[current_index],
        "--auth-user-pass", auth_file,
    ])
    sleep(10)  # Wait for connection


search_request = "Samsung"

limit = 300  # Items limit

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()

    page.goto("https://www.amazon.com", timeout=60000)


    b = page.query_selector('button[alt="Continue shopping"]')
    if b:
        b.click()

    page.wait_for_selector("input#twotabsearchtextbox")
    page.fill("input#twotabsearchtextbox", search_request)
    page.press("input#twotabsearchtextbox", "Enter")
    page.wait_for_load_state("load")

    wb = Workbook()
    ws = wb.active
    ws.title = "amazon listing of products"

    headers = ["Name", "Product URL", "Price", "Rating", "Number of sellers", "Number of reviews"]
    ws.append(headers)
    ws['A1'].font = ws['B1'].font = ws['C1'].font = ws['D1'].font = ws['E1'].font = ws['F1'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 50
    for row in ws.iter_rows(min_col=2, max_col=2, min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    pp = context.new_page()

    while True:
        if limit and ws.max_row - 1 >= limit:  # subtract header row
            break

        for el in page.query_selector_all('div[role="listitem"]'):
            link_el = el.query_selector("a")
            if not link_el:
                break

            href = link_el.get_attribute("href")
            if not href:
                print("no href { ", el.inner_html(), " }")
                continue

            pp.goto(f"https://www.amazon.com{href}")
            pp.wait_for_load_state("load")
            sleep(random.uniform(1, 2))

            name_el = pp.query_selector('span[id="productTitle"]')
            name = name_el.inner_text().strip() if name_el else "N/A"

            price_el = pp.query_selector("#ppd .a-price span:nth-child(2)")
            if not price_el or price_el.text_content().strip() == "":
              price_el = pp.query_selector("#ppd > #usedBuySection > div.a-row.a-grid-vertical-align.a-grid-center "
                                           "> div > span.a-size-base.a-color-price.offer-price.a-text-normal")
            pricestr = price_el.inner_text().strip().replace('\n', '') if price_el else "N/A"
            price = pricestr if pricestr != "N/A" else "N/A"
            if price == "N/A":
                print("no price in ", pp.url)
                wb.save("sample_data.xlsx")
            rating = pp.query_selector("#acrPopover")
            rating = re.match(r"(\d+\.\d+)",
                              rating.text_content().strip()).group(1) if rating else "N/A"
            nos = 1  # number of sellers
            a = pp.query_selector('a[id="aod-ingress-link"]')
            if a:
                a.click()
                h5 = 'h5[id="aod-filter-offer-count-string"]'
                try:
                    pp.wait_for_selector(h5)
                    text = pp.query_selector(h5).text_content()
                    count = int(re.search(r'\d+', text).group())
                    nos += count
                except Exception:
                    pass
            nof = pp.query_selector("#acrCustomerReviewText")
            nof = re.search(r"(\d+)", nof.text_content().strip()).group(1) if nof else "N/A"
            ws.append([name, f"https://www.amazon.com{href}", price, rating, nos, nof])

        next_link = page.locator('span.s-pagination-strip > ul > li').last.locator("span > a")
        if next_link.is_visible():
            sleep(random.uniform(4, 8))
            next_link.click()
            sleep(random.uniform(4, 8))
        else:
            break

    wb.save("sample_data.xlsx")